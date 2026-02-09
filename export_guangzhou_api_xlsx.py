#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 Open-Meteo API 原始日序列数据导出为 xlsx（元数据）。
与报告使用的数据源一致：广州 2016-01-01 至 2025-12-31 逐日气温。
"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 复用报告的拉取与解析逻辑
try:
    from guangzhou_weather_report import fetch_data, to_dataframe, API_URL, PARAMS
except ImportError:
    fetch_data = to_dataframe = None
    API_URL = "https://archive-api.open-meteo.com/v1/archive"
    PARAMS = {
        "latitude": 23.1291,
        "longitude": 113.2644,
        "start_date": "2016-01-01",
        "end_date": "2025-12-31",
        "daily": "temperature_2m_max,temperature_2m_min,temperature_2m_mean",
        "timezone": "Asia/Shanghai",
    }
    import json
    import urllib.request
    import urllib.parse
    import pandas as pd

    def fetch_data():
        query = urllib.parse.urlencode(PARAMS)
        with urllib.request.urlopen(f"{API_URL}?{query}") as resp:
            return json.loads(resp.read().decode())

    def to_dataframe(raw):
        daily = raw.get("daily", {})
        df = pd.DataFrame({
            "date": pd.to_datetime(daily.get("time", [])),
            "temp_max": daily.get("temperature_2m_max", []),
            "temp_min": daily.get("temperature_2m_min", []),
            "temp_mean": daily.get("temperature_2m_mean", []),
        })
        df["year"] = df["date"].dt.year
        df["month"] = df["date"].dt.month
        df["temp_range"] = df["temp_max"] - df["temp_min"]
        return df


def _apply_sheet_style(ws, font_name="Arial", font_size=11):
    header_font = Font(name=font_name, size=font_size, bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name=font_name, size=font_size)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    for col in range(1, ws.max_column + 1):
        c = get_column_letter(col)
        w = ws.column_dimensions[c].width or 10
        ws.column_dimensions[c].width = max(10, min(18, w))


def export_api_to_xlsx(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "API日序列"
    ws.append(["日期", "最高温(°C)", "最低温(°C)", "平均温(°C)", "年", "月", "日温差(°C)"])
    for _, row in df.sort_values("date").iterrows():
        ws.append([
            row["date"].strftime("%Y-%m-%d"),
            round(float(row["temp_max"]), 1),
            round(float(row["temp_min"]), 1),
            round(float(row["temp_mean"]), 1),
            int(row["year"]),
            int(row["month"]),
            round(float(row["temp_range"]), 1),
        ])
    _apply_sheet_style(ws)
    wb.save(out_path)


def main():
    if not HAS_OPENPYXL:
        print("请先安装 openpyxl：pip install openpyxl")
        sys.exit(1)
    if fetch_data is None or to_dataframe is None:
        print("无法加载 API 拉取逻辑，请在本项目目录运行并确保 guangzhou_weather_report 可用。")
        sys.exit(1)
    print("正在从 Open-Meteo API 获取数据...")
    raw = fetch_data()
    df = to_dataframe(raw)
    print(f"已加载 {len(df)} 行日序列数据")
    base = Path(__file__).parent
    out_path = base / "广州天气_API元数据.xlsx"
    export_api_to_xlsx(df, out_path)
    print(f"已导出：{out_path}")


if __name__ == "__main__":
    main()
