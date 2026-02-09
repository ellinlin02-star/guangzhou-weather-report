#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将广州天气报告分析结果导出为 xlsx。
数据来源：guangzhou_weather_report_data.json（与报告页「表格数据」一致）
"""

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_report_data(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _chart_data(report):
    """与 guangzhou_weather_report.py 中 _chart_data 一致，从 report 提取图表用表数据"""
    y25 = report["year_2025"]
    y10 = report["years_10"]
    ab = report["abnormal"]
    pred = report["predict_2026"]
    months_cn = "1月,2月,3月,4月,5月,6月,7月,8月,9月,10月,11月,12月".split(",")

    chart_2025_daily = {"labels": [], "max": [], "min": [], "mean": []}
    if y25.get("daily"):
        chart_2025_daily["labels"] = [x["date"] for x in y25["daily"]]
        chart_2025_daily["max"] = [float(x["temp_max"]) for x in y25["daily"]]
        chart_2025_daily["min"] = [float(x["temp_min"]) for x in y25["daily"]]
        chart_2025_daily["mean"] = [float(x["temp_mean"]) for x in y25["daily"]]

    chart_2025 = {"labels": months_cn}
    if y25.get("by_month"):
        chart_2025["max"] = [float(y25["by_month"][str(m)]["temp_max_mean"]) for m in range(1, 13)]
        chart_2025["min"] = [float(y25["by_month"][str(m)]["temp_min_mean"]) for m in range(1, 13)]
        chart_2025["mean"] = [float(y25["by_month"][str(m)]["temp_mean_mean"]) for m in range(1, 13)]

    by_year = sorted(y10["by_year"].items())
    chart_10y = {
        "labels": [yr for yr, _ in by_year],
        "max": [float(v["temp_max_mean"]) for _, v in by_year],
        "min": [float(v["temp_min_mean"]) for _, v in by_year],
        "mean": [float(v["temp_mean_mean"]) for _, v in by_year],
    }

    ab_by = ab.get("by_year", {})
    by_year_ab = sorted(ab_by.items())
    chart_ab = {
        "labels": [yr for yr, _ in by_year_ab],
        "high": [v["high"] for _, v in by_year_ab],
        "low": [v["low"] for _, v in by_year_ab],
    }

    chart_2026 = {
        "labels": months_cn,
        "mean": [pred["by_month"][str(m)]["temp_mean"] for m in range(1, 13)],
        "max": [pred["by_month"][str(m)]["temp_max"] for m in range(1, 13)],
        "min": [pred["by_month"][str(m)]["temp_min"] for m in range(1, 13)],
    }
    return {
        "chart_2025_daily": chart_2025_daily,
        "chart_2025": chart_2025,
        "chart_10y": chart_10y,
        "chart_ab": chart_ab,
        "chart_2026": chart_2026,
    }


def _apply_sheet_style(ws, font_name="Arial", font_size=11):
    """统一字体与表头样式（符合 xlsx skill：专业字体、无公式错误）"""
    header_font = Font(name=font_name, size=font_size, bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name=font_name, size=font_size)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(20, ws.column_dimensions[get_column_letter(col)].width or 10))


def export_to_xlsx(report, out_path):
    wb = Workbook()
    # 删除默认 sheet，后面按顺序添加
    wb.remove(wb.active)
    data = _chart_data(report)

    # 1. 2025 年逐日气温
    d = data["chart_2025_daily"]
    if d.get("labels"):
        ws = wb.create_sheet("2025逐日", 0)
        ws.append(["日期", "日最高温(°C)", "日平均温(°C)", "日最低温(°C)"])
        for i in range(len(d["labels"])):
            ws.append([d["labels"][i], d["max"][i], d["mean"][i], d["min"][i]])
        _apply_sheet_style(ws)

    # 2. 2025 年月度气温
    d = data["chart_2025"]
    if d.get("labels"):
        ws = wb.create_sheet("2025月度", 1)
        ws.append(["月份", "月均最高温(°C)", "月均平均温(°C)", "月均最低温(°C)"])
        for i in range(len(d["labels"])):
            ws.append([d["labels"][i], d["max"][i], d["mean"][i], d["min"][i]])
        _apply_sheet_style(ws)

    # 3. 近10年气温趋势
    d = data["chart_10y"]
    if d.get("labels"):
        ws = wb.create_sheet("近10年", 2)
        ws.append(["年份", "年均最高温(°C)", "年均平均温(°C)", "年均最低温(°C)"])
        for i in range(len(d["labels"])):
            ws.append([d["labels"][i], d["max"][i], d["mean"][i], d["min"][i]])
        _apply_sheet_style(ws)

    # 4. 异常天气
    d = data["chart_ab"]
    if d.get("labels"):
        ws = wb.create_sheet("异常天气", 3)
        ws.append(["年份", "高温异常(天)", "低温异常(天)"])
        for i in range(len(d["labels"])):
            ws.append([d["labels"][i], d["high"][i], d["low"][i]])
        _apply_sheet_style(ws)

    # 5. 2026 年预测
    d = data["chart_2026"]
    if d.get("labels"):
        ws = wb.create_sheet("2026预测", 4)
        ws.append(["月份", "预测最高温(°C)", "预测平均温(°C)", "预测最低温(°C)"])
        for i in range(len(d["labels"])):
            ws.append([d["labels"][i], d["max"][i], d["mean"][i], d["min"][i]])
        _apply_sheet_style(ws)

    wb.save(out_path)


def main():
    if not HAS_OPENPYXL:
        print("请先安装 openpyxl：pip install openpyxl")
        return
    base = Path(__file__).parent
    json_path = base / "guangzhou_weather_report_data.json"
    if not json_path.exists():
        print(f"未找到数据文件：{json_path}，请先运行 guangzhou_weather_report.py 生成报告数据。")
        return
    report = load_report_data(json_path)
    out_path = base / "广州天气报告_分析结果.xlsx"
    export_to_xlsx(report, out_path)
    print(f"已导出：{out_path}")


if __name__ == "__main__":
    main()
